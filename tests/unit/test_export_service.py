"""Tests for the ExportService class."""

import pytest
import json
from io import BytesIO
from openpyxl import load_workbook

from src.services.export_service import ExportService
from src.exceptions import ExportError, ValidationError


class TestExportService:
    """Test cases for ExportService."""
    
    def setup_method(self):
        """Set up test fixtures."""
        self.export_service = ExportService()
        
        # Sample test data
        self.sample_tasks = [
            {
                'summary': 'Test Task 1',
                'description': 'Description for task 1',
                'issue_type': 'Task',
                'reporter': 'test@example.com',
                'due_date': '2024-12-31'
            },
            {
                'summary': 'Test Task 2',
                'description': 'Description for task 2',
                'issue_type': 'Bug',
                'reporter': 'test@example.com',
                'due_date': ''
            }
        ]
        
        self.sample_qa_items = [
            {
                'question': 'What is the deadline?',
                'answer': 'End of December',
                'asked_by': 'user@example.com',
                'answered_by': 'manager@example.com',
                'status': 'answered',
                'context': 'Project planning meeting'
            },
            {
                'question': 'Who will handle testing?',
                'answer': '',
                'asked_by': 'user@example.com',
                'answered_by': '',
                'status': 'unanswered',
                'context': 'Development discussion'
            }
        ]
    
    def test_export_csv_standard_template(self):
        """Test CSV export with standard template."""
        content_bytes, filename, mimetype = self.export_service.export_data(
            tasks=self.sample_tasks,
            export_format='csv',
            template='standard'
        )
        
        assert isinstance(content_bytes, bytes)
        assert filename.endswith('.csv')
        assert mimetype == 'text/csv'
        
        # Check content
        content = content_bytes.decode('utf-8')
        assert 'Summary' in content and 'Description' in content and 'Issue Type' in content
        assert 'Test Task 1' in content
        assert 'Test Task 2' in content
    
    def test_export_csv_summary_template(self):
        """Test CSV export with summary template."""
        content_bytes, filename, mimetype = self.export_service.export_data(
            tasks=self.sample_tasks,
            export_format='csv',
            template='summary'
        )
        
        content = content_bytes.decode('utf-8')
        assert 'Summary' in content and 'Issue Type' in content and 'Due Date' in content
        assert 'Description' not in content  # Should not include description in summary
        assert 'Test Task 1' in content
    
    def test_export_csv_detailed_template(self):
        """Test CSV export with detailed template including Q&A."""
        content_bytes, filename, mimetype = self.export_service.export_data(
            tasks=self.sample_tasks,
            qa_items=self.sample_qa_items,
            export_format='csv',
            template='detailed'
        )
        
        content = content_bytes.decode('utf-8')
        assert '=== TASKS ===' in content
        assert '=== Q&A ITEMS ===' in content
        assert 'What is the deadline?' in content
        assert 'Test Task 1' in content
    
    def test_export_json_standard_template(self):
        """Test JSON export with standard template."""
        content_bytes, filename, mimetype = self.export_service.export_data(
            tasks=self.sample_tasks,
            export_format='json',
            template='standard'
        )
        
        assert filename.endswith('.json')
        assert mimetype == 'application/json'
        
        # Parse and check JSON content
        content = json.loads(content_bytes.decode('utf-8'))
        assert 'export_info' in content
        assert 'tasks' in content
        assert len(content['tasks']) == 2
        assert content['tasks'][0]['summary'] == 'Test Task 1'
    
    def test_export_json_detailed_template(self):
        """Test JSON export with detailed template including statistics."""
        content_bytes, filename, mimetype = self.export_service.export_data(
            tasks=self.sample_tasks,
            qa_items=self.sample_qa_items,
            export_format='json',
            template='detailed'
        )
        
        content = json.loads(content_bytes.decode('utf-8'))
        assert 'statistics' in content
        assert content['statistics']['total_tasks'] == 2
        assert content['statistics']['total_qa_items'] == 2
        assert content['statistics']['answered_questions'] == 1
    
    def test_export_excel_standard_template(self):
        """Test Excel export with standard template."""
        content_bytes, filename, mimetype = self.export_service.export_data(
            tasks=self.sample_tasks,
            export_format='excel',
            template='standard'
        )
        
        assert filename.endswith('.xlsx')
        assert mimetype == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        # Load and check Excel content
        wb = load_workbook(BytesIO(content_bytes))
        ws = wb.active
        assert ws.title == 'JIRA Tasks'
        assert ws['A1'].value == 'Summary'
        assert ws['A2'].value == 'Test Task 1'
    
    def test_export_excel_detailed_template(self):
        """Test Excel export with detailed template (multiple sheets)."""
        content_bytes, filename, mimetype = self.export_service.export_data(
            tasks=self.sample_tasks,
            qa_items=self.sample_qa_items,
            export_format='excel',
            template='detailed'
        )
        
        wb = load_workbook(BytesIO(content_bytes))
        sheet_names = wb.sheetnames
        assert 'Tasks' in sheet_names
        assert 'Q&A Items' in sheet_names
        
        # Check tasks sheet
        tasks_sheet = wb['Tasks']
        assert tasks_sheet['A1'].value == 'Summary'
        assert tasks_sheet['A2'].value == 'Test Task 1'
        
        # Check Q&A sheet
        qa_sheet = wb['Q&A Items']
        assert qa_sheet['A1'].value == 'Question'
        assert qa_sheet['A2'].value == 'What is the deadline?'
    
    def test_export_with_no_data_raises_error(self):
        """Test that export with no data raises ValidationError."""
        with pytest.raises(ValidationError, match="No data provided for export"):
            self.export_service.export_data()
    
    def test_export_with_invalid_format_raises_error(self):
        """Test that export with invalid format raises ValidationError."""
        with pytest.raises(ValidationError, match="Unsupported export format"):
            self.export_service.export_data(
                tasks=self.sample_tasks,
                export_format='invalid_format'
            )
    
    def test_export_with_invalid_task_data(self):
        """Test export with invalid task data (should skip invalid tasks)."""
        invalid_tasks = [
            {'summary': ''},  # Invalid: empty summary
            {'summary': 'Valid Task', 'issue_type': 'Task'}  # Valid
        ]
        
        content_bytes, filename, mimetype = self.export_service.export_data(
            tasks=invalid_tasks,
            export_format='csv',
            template='standard'
        )
        
        content = content_bytes.decode('utf-8')
        assert 'Valid Task' in content
        # Should only have header + 1 valid task
        lines = content.strip().split('\n')
        assert len(lines) == 2  # Header + 1 data row
    
    def test_get_supported_formats(self):
        """Test getting supported export formats."""
        formats = self.export_service.get_supported_formats()
        
        assert 'csv' in formats
        assert 'json' in formats
        assert 'excel' in formats
        assert formats['csv']['extension'] == 'csv'
        assert formats['json']['mimetype'] == 'application/json'
    
    def test_get_supported_templates(self):
        """Test getting supported export templates."""
        templates = self.export_service.get_supported_templates()
        
        assert 'standard' in templates
        assert 'summary' in templates
        assert 'detailed' in templates
        assert 'Standard format' in templates['standard']
    
    def test_filename_generation(self):
        """Test filename generation for different formats and templates."""
        # Test CSV filename
        content_bytes, filename, mimetype = self.export_service.export_data(
            tasks=self.sample_tasks,
            export_format='csv',
            template='summary'
        )
        assert 'jira_summary_' in filename
        assert filename.endswith('.csv')
        
        # Test Excel filename
        content_bytes, filename, mimetype = self.export_service.export_data(
            tasks=self.sample_tasks,
            export_format='excel',
            template='detailed'
        )
        assert 'jira_detailed_' in filename
        assert filename.endswith('.xlsx')
    
    def test_export_qa_items_only(self):
        """Test export with only Q&A items (no tasks)."""
        content_bytes, filename, mimetype = self.export_service.export_data(
            qa_items=self.sample_qa_items,
            export_format='json',
            template='detailed'
        )
        
        content = json.loads(content_bytes.decode('utf-8'))
        assert 'qa_items' in content
        assert len(content['qa_items']) == 2
        assert content['statistics']['total_tasks'] == 0
        assert content['statistics']['total_qa_items'] == 2
