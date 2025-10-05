"""Enhanced export service supporting multiple formats (CSV, JSON, Excel)."""

import csv
import json
import io
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..models.task import JiraTask
from ..models.qa_item import QAItem
from ..exceptions import ExportError, ValidationError
from ..utils import LoggerMixin


class ExportService(LoggerMixin):
    """Enhanced export service supporting CSV, JSON, and Excel formats."""
    
    def __init__(self):
        """Initialize export service."""
        self.task_headers = ['Summary', 'Description', 'Issue Type']
        self.qa_headers = ['Question', 'Answer', 'Asked By', 'Answered By', 'Status', 'Context']
        
        # Export format configurations
        self.export_formats = {
            'csv': {
                'extension': 'csv',
                'mimetype': 'text/csv',
                'description': 'Comma-separated values format'
            },
            'json': {
                'extension': 'json',
                'mimetype': 'application/json',
                'description': 'JavaScript Object Notation format'
            },
            'excel': {
                'extension': 'xlsx',
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'description': 'Microsoft Excel format'
            }
        }
    
    def export_data(self, 
                   tasks: Optional[List[Dict[str, Any]]] = None,
                   qa_items: Optional[List[Dict[str, Any]]] = None,
                   export_format: str = 'csv',
                   template: str = 'standard') -> Tuple[bytes, str, str]:
        """
        Export tasks and Q&A items in the specified format.
        
        Args:
            tasks: List of task dictionaries
            qa_items: List of Q&A item dictionaries
            export_format: Export format ('csv', 'json', 'excel')
            template: Export template ('standard', 'detailed', 'summary')
            
        Returns:
            Tuple of (content_bytes, filename, mimetype)
            
        Raises:
            ExportError: If export fails
            ValidationError: If data is invalid
        """
        if not tasks and not qa_items:
            raise ValidationError("No data provided for export")
        
        if export_format not in self.export_formats:
            raise ValidationError(f"Unsupported export format: {export_format}")
        
        try:
            self.logger.info(f"Exporting data in {export_format} format with {template} template")
            
            # Validate and convert data
            validated_tasks = self._validate_tasks(tasks or [])
            validated_qa_items = self._validate_qa_items(qa_items or [])
            
            # Generate content based on format
            if export_format == 'csv':
                content = self._export_csv(validated_tasks, validated_qa_items, template)
                content_bytes = content.encode('utf-8')
            elif export_format == 'json':
                content = self._export_json(validated_tasks, validated_qa_items, template)
                content_bytes = content.encode('utf-8')
            elif export_format == 'excel':
                content_bytes = self._export_excel(validated_tasks, validated_qa_items, template)
            
            # Generate filename
            filename = self._generate_filename(export_format, template)
            mimetype = self.export_formats[export_format]['mimetype']
            
            self.logger.info(f"Export completed successfully: {filename}")
            return content_bytes, filename, mimetype
            
        except Exception as e:
            self.logger.error(f"Export failed: {e}")
            raise ExportError(f"Failed to export data: {str(e)}")
    
    def _validate_tasks(self, tasks: List[Dict[str, Any]]) -> List[JiraTask]:
        """Validate and convert task dictionaries to JiraTask objects."""
        validated_tasks = []

        for i, task_data in enumerate(tasks):
            try:
                jira_task = JiraTask(
                    summary=task_data.get('summary', ''),
                    description=task_data.get('description', ''),
                    issue_type=task_data.get('issue_type', 'Task')
                )
                validated_tasks.append(jira_task)

            except ValueError as e:
                self.logger.warning(f"Skipping invalid task {i+1}: {e}")
                continue

        return validated_tasks
    
    def _validate_qa_items(self, qa_items: List[Dict[str, Any]]) -> List[QAItem]:
        """Validate and convert Q&A dictionaries to QAItem objects."""
        validated_qa_items = []
        
        for i, qa_data in enumerate(qa_items):
            try:
                qa_item = QAItem(
                    question=qa_data.get('question', ''),
                    answer=qa_data.get('answer', ''),
                    asked_by=qa_data.get('asked_by', 'meeting@example.com'),
                    answered_by=qa_data.get('answered_by', ''),
                    context=qa_data.get('context', ''),
                    status=qa_data.get('status', 'unanswered')
                )
                validated_qa_items.append(qa_item)
                
            except ValueError as e:
                self.logger.warning(f"Skipping invalid Q&A item {i+1}: {e}")
                continue
        
        return validated_qa_items
    
    def _export_csv(self, tasks: List[JiraTask], qa_items: List[QAItem], template: str) -> str:
        """Export data as CSV format."""
        output = io.StringIO()
        
        if template == 'summary':
            # Summary template: tasks only, minimal columns
            writer = csv.writer(output, quoting=csv.QUOTE_ALL)
            writer.writerow(['Summary', 'Issue Type'])

            for task in tasks:
                writer.writerow([task.summary, task.issue_type])
        
        elif template == 'detailed':
            # Detailed template: separate sections for tasks and Q&A
            writer = csv.writer(output, quoting=csv.QUOTE_ALL)
            
            # Tasks section
            if tasks:
                writer.writerow(['=== TASKS ==='])
                writer.writerow(self.task_headers)
                for task in tasks:
                    writer.writerow([
                        task.summary, task.description, task.issue_type
                    ])
                writer.writerow([])  # Empty row separator
            
            # Q&A section
            if qa_items:
                writer.writerow(['=== Q&A ITEMS ==='])
                writer.writerow(self.qa_headers)
                for qa in qa_items:
                    writer.writerow([
                        qa.question, qa.answer, qa.asked_by,
                        qa.answered_by, qa.status, qa.context
                    ])
        
        else:  # standard template
            # Standard template: tasks only, all columns
            writer = csv.writer(output, quoting=csv.QUOTE_ALL)
            writer.writerow(self.task_headers)
            
            for task in tasks:
                writer.writerow([
                    task.summary, task.description, task.issue_type
                ])
        
        content = output.getvalue()
        output.close()
        return content
    
    def _export_json(self, tasks: List[JiraTask], qa_items: List[QAItem], template: str) -> str:
        """Export data as JSON format."""
        data = {
            'export_info': {
                'timestamp': datetime.now().isoformat(),
                'template': template,
                'format': 'json'
            }
        }
        
        if template == 'summary':
            # Summary template: minimal task data
            data['tasks'] = [
                {
                    'summary': task.summary,
                    'issue_type': task.issue_type,
                    'due_date': task.due_date
                }
                for task in tasks
            ]
        
        elif template == 'detailed':
            # Detailed template: full data with metadata
            data['tasks'] = [task.to_dict() for task in tasks]
            data['qa_items'] = [qa.to_dict() for qa in qa_items]
            data['statistics'] = {
                'total_tasks': len(tasks),
                'total_qa_items': len(qa_items),
                'task_types': list(set(task.issue_type for task in tasks)),
                'answered_questions': len([qa for qa in qa_items if qa.has_answer()])
            }
        
        else:  # standard template
            # Standard template: tasks with basic structure
            data['tasks'] = [task.to_dict() for task in tasks]
            if qa_items:
                data['qa_items'] = [qa.to_dict() for qa in qa_items]
        
        return json.dumps(data, indent=2, ensure_ascii=False)
    
    def _export_excel(self, tasks: List[JiraTask], qa_items: List[QAItem], template: str) -> bytes:
        """Export data as Excel format with formatting."""
        wb = Workbook()
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        if template == 'summary':
            # Summary template: single sheet with minimal task data
            ws = wb.active
            ws.title = "Task Summary"
            
            # Headers
            headers = ['Summary', 'Issue Type']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            # Data
            for row, task in enumerate(tasks, 2):
                ws.cell(row=row, column=1, value=task.summary).border = border
                ws.cell(row=row, column=2, value=task.issue_type).border = border

            # Auto-adjust column widths
            for col in range(1, 3):
                ws.column_dimensions[get_column_letter(col)].width = 20
        
        elif template == 'detailed':
            # Detailed template: separate sheets for tasks and Q&A
            
            # Tasks sheet
            if tasks:
                ws_tasks = wb.active
                ws_tasks.title = "Tasks"
                
                # Headers
                for col, header in enumerate(self.task_headers, 1):
                    cell = ws_tasks.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                
                # Data
                for row, task in enumerate(tasks, 2):
                    ws_tasks.cell(row=row, column=1, value=task.summary).border = border
                    ws_tasks.cell(row=row, column=2, value=task.description).border = border
                    ws_tasks.cell(row=row, column=3, value=task.issue_type).border = border

                # Auto-adjust column widths
                for col in range(1, 4):
                    ws_tasks.column_dimensions[get_column_letter(col)].width = 25
            
            # Q&A sheet
            if qa_items:
                ws_qa = wb.create_sheet("Q&A Items")
                
                # Headers
                for col, header in enumerate(self.qa_headers, 1):
                    cell = ws_qa.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                
                # Data
                for row, qa in enumerate(qa_items, 2):
                    ws_qa.cell(row=row, column=1, value=qa.question).border = border
                    ws_qa.cell(row=row, column=2, value=qa.answer).border = border
                    ws_qa.cell(row=row, column=3, value=qa.asked_by).border = border
                    ws_qa.cell(row=row, column=4, value=qa.answered_by).border = border
                    ws_qa.cell(row=row, column=5, value=qa.status).border = border
                    ws_qa.cell(row=row, column=6, value=qa.context).border = border
                
                # Auto-adjust column widths
                for col in range(1, 7):
                    ws_qa.column_dimensions[get_column_letter(col)].width = 20
        
        else:  # standard template
            # Standard template: tasks with full formatting
            ws = wb.active
            ws.title = "JIRA Tasks"
            
            # Headers
            for col, header in enumerate(self.task_headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Data
            for row, task in enumerate(tasks, 2):
                ws.cell(row=row, column=1, value=task.summary).border = border
                ws.cell(row=row, column=2, value=task.description).border = border
                ws.cell(row=row, column=3, value=task.issue_type).border = border

            # Auto-adjust column widths
            for col in range(1, 4):
                ws.column_dimensions[get_column_letter(col)].width = 25
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def _generate_filename(self, export_format: str, template: str) -> str:
        """Generate filename for export."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        extension = self.export_formats[export_format]['extension']
        
        if template == 'summary':
            prefix = "jira_summary"
        elif template == 'detailed':
            prefix = "jira_detailed"
        else:
            prefix = "jira_tasks"
        
        return f"{prefix}_{timestamp}.{extension}"
    
    def get_supported_formats(self) -> Dict[str, Dict[str, str]]:
        """Get list of supported export formats."""
        return self.export_formats.copy()
    
    def get_supported_templates(self) -> Dict[str, str]:
        """Get list of supported export templates."""
        return {
            'standard': 'Standard format with all task fields',
            'detailed': 'Detailed format with tasks and Q&A items',
            'summary': 'Summary format with essential fields only'
        }
